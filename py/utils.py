import os
import re
import shutil
import sys
import threading
import time
import fitz
import openpyxl
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterator, TypedDict

# ═══════════════════════════════════════════════════════════════════
#  Constants
# ═══════════════════════════════════════════════════════════════════

POLICY_FLAGS: list[tuple[str, str]] = [
    ("top", "Top"),
    ("storage", "Storage"),
    ("cancellation", "Cancel"),
    ("rental", "Rental"),
    ("special_risk", "Special Risk"),
    ("garage", "Garage"),
    ("manuscript", "Manuscript"),
    ("binder", "Binder"),
]

_NON_DISPLAY_PLATES = frozenset({"NONLIC", "STORAGE", "DEALER"})
_REGISTRATION_PLATES = frozenset({"NONLIC", "DEALER"})

_RE_INVALID = re.compile(r'[.:/\\*?"<>|]')
_RE_SPACES = re.compile(r"\s+")
_RE_COUNTER = re.compile(r"\s*\(\d+\)$")
_RE_COMPANY = re.compile(r"(Inc\.?|Ltd\.?|Corp\.?)$", re.IGNORECASE)
_RE_YEAR = re.compile(r"^\d{4}$")
_RE_FILENAME_TS = re.compile(r"\[([^\]]+)\]")

# ═══════════════════════════════════════════════════════════════════
#  ICBC Patterns & Page Rects
# ═══════════════════════════════════════════════════════════════════

ICBC_PATTERNS: "RegexPatterns" = {
    "timestamp": re.compile(r"Transaction Timestamp\s*(\d{14})"),
    "certificate_replacement": re.compile(r"Certificate Replacement\s*(\d{14})"),
    "same_day_re-print": re.compile(r"Same day Re-print\s*(\d{14})"),
    "license_plate": re.compile(
        r"Licence Plate Number\s*([A-Z0-9\- ]+)", re.IGNORECASE
    ),
    "temporary_operation_permit": re.compile(
        r"Temporary Operation Permit and Owner['\u2019]s Certificate of Insurance",
        re.IGNORECASE,
    ),
    "agency_number": re.compile(r"Agency Number\s*[:#]?\s*(\d{1,6})", re.IGNORECASE),
    "customer_copy": re.compile(r"customer copy", re.IGNORECASE),
    "validation_stamp": re.compile(r"NOT VALID UNLESS STAMPED BY", re.IGNORECASE),
    "time_of_validation": re.compile(r"TIME OF VALIDATION", re.IGNORECASE),
    "producer": re.compile(r"-\s*([A-Za-z]+)\s*-", re.IGNORECASE),
    "transaction_type": re.compile(r"Transaction Type\s+([A-Z]+)", re.IGNORECASE),
    "cancellation": re.compile(r"Application for Cancellation"),
    "storage_policy": re.compile(r"Storage Policy"),
    "rental_vehicle_policy": re.compile(r"Rental Vehicle Policy"),
    "special_risk_own_damage_policy": re.compile(r"Special Risk Own Damage Policy"),
    "garage_vehicle_certificate": re.compile(r"Garage Vehicle Certificate"),
    "payment_plan": re.compile(r"Payment Plan Agreement"),
    "payment_plan_receipt": re.compile(r"Payment Plan Receipt"),
    "manuscript": re.compile(r"Manuscript Certificate/Manuscript Policy"),
    "binder": re.compile(
        r"Binder for Owner['\u2019]s Interim Certificate of Insurance"
    ),
    "has_bcdl": re.compile(
        r"Owner['\u2019]s BC Driver['\u2019]s Licence Number(?:\s+(\*{4,5}\d{3}))?",
        re.IGNORECASE,
    ),
}

PAGE_RECTS: "PageRects" = {
    "timestamp": fitz.Rect(409.979, 63.8488, 576.0, 83.7455),
    "producer": fitz.Rect(198.0, 752.729736328125, 255.011, 769.977),
    "customer_copy": fitz.Rect(498.438, 751.953, 578.181, 769.977),
}


# ═══════════════════════════════════════════════════════════════════
#  Type Definitions
# ═══════════════════════════════════════════════════════════════════


class RegexPatterns(TypedDict, total=False):
    timestamp: re.Pattern[str]
    certificate_replacement: re.Pattern[str]
    same_day_reprint: re.Pattern[str]
    license_plate: re.Pattern[str]
    has_bcdl: re.Pattern[str]
    temporary_operation_permit: re.Pattern[str]
    payment_plan: re.Pattern[str]
    payment_plan_receipt: re.Pattern[str]
    agency_number: re.Pattern[str]
    customer_copy: re.Pattern[str]
    validation_stamp: re.Pattern[str]
    time_of_validation: re.Pattern[str]
    producer: re.Pattern[str]
    transaction_type: re.Pattern[str]
    storage_policy: re.Pattern[str]
    cancellation: re.Pattern[str]
    special_risk_own_damage_policy: re.Pattern[str]
    rental_vehicle_policy: re.Pattern[str]
    garage_vehicle_certificate: re.Pattern[str]
    manuscript: re.Pattern[str]
    binder: re.Pattern[str]


class PageRects(TypedDict, total=False):
    timestamp: fitz.Rect
    producer: fitz.Rect


# ═══════════════════════════════════════════════════════════════════
#  Data Models
# ═══════════════════════════════════════════════════════════════════


@dataclass
class FolderMapping:
    tool_event: str | None
    copy_input_folder: Path | None
    create_folder_tool_output_folder: Path | None
    e_stamp_output_folder: Path | None
    agency_number: str | None = None
    producer_mapping: dict[str, str] = field(default_factory=dict)


@dataclass
class ScanResult:
    documents: dict[Path, "ICBCDocument"]
    non_icbc: list[Path]
    payment_plans: list[Path]
    unreadable: list[Path]


@dataclass
class ICBCDocument:
    path: Path
    transaction_timestamp: str
    certificate_replacement: str | None = None
    same_day_reprint: str | None = None
    license_plate: str | None = None
    insured_name: str | None = None
    producer_name: str | None = None
    transaction_type: str | None = None
    # policy flags
    top: bool = False
    storage: bool = False
    cancellation: bool = False
    special_risk: bool = False
    rental: bool = False
    garage: bool = False
    manuscript: bool = False
    binder: bool = False
    # stamping-mode fields (populated only when stamping_mode=True)
    agency_number: str | None = None
    customer_copy_pages: list[int] = field(default_factory=list)
    validation_stamp_coords: list[tuple] = field(default_factory=list)
    time_of_validation_coords: list[tuple] = field(default_factory=list)

    # ── Properties ──────────────────────────────────────────────── #

    @property
    def plate(self) -> str:
        return (self.license_plate or "").strip().upper()

    @property
    def clean_name(self) -> str:
        return _sanitise(self.insured_name or "").title()

    @property
    def name_prefix(self) -> str:
        return self.base_name().split(" - ", 1)[0].strip()

    # ── Name builders ────────────────────────────────────────────── #

    def base_name(self) -> str:
        return self._build_name(include_change_cancel=True)

    def stamp_name(self) -> str:
        return self._build_name(include_change_cancel=False)

    def _build_name(self, *, include_change_cancel: bool) -> str:
        if self.plate and self.plate not in _NON_DISPLAY_PLATES:
            core = (
                f"{self.clean_name} - {self.plate}"
                if include_change_cancel
                else self.plate
            )
        elif self.clean_name:
            core = self.clean_name
        elif self.transaction_timestamp:
            core = self.transaction_timestamp
        else:
            core = "UNKNOWN"

        return self._apply_suffix(core, include_change_cancel=include_change_cancel)

    def _apply_suffix(self, core: str, *, include_change_cancel: bool) -> str:
        for attr, label in POLICY_FLAGS:
            if not getattr(self, attr, False):
                continue
            if label == "Cancel" and not include_change_cancel:
                break
            return f"{core} - {label}" if label != "Cancel" else f"{core} {label}"

        if (
            include_change_cancel
            and (self.transaction_type or "").strip().title() == "Change"
        ):
            return f"{core} Change"

        if self.plate in _REGISTRATION_PLATES:
            return f"{core} - Registration"

        return core


# ═══════════════════════════════════════════════════════════════════
#  Progress Bar
# ═══════════════════════════════════════════════════════════════════


def progressbar(
    it,
    prefix: str = "",
    size: int = 60,
    out=sys.stdout,
    count: int | None = None,
) -> Iterator:
    total = count if count is not None else len(it)  # type: ignore[arg-type]
    start = time.time()

    def _render(j: int | float) -> None:
        filled = int(size * j / total)
        elapsed = time.time() - start
        remaining = (elapsed / j) * (total - j) if j else 0
        mins, secs = divmod(remaining, 60)
        print(
            f"{prefix}[{'█' * filled}{'.' * (size - filled)}] "
            f"{j}/{total} Est wait {int(mins):02}:{secs:03.1f}",
            end="\r",
            file=out,
            flush=True,
        )

    if total > 0:
        _render(0.1)
        for i, item in enumerate(it):
            yield item
            _render(i + 1)
        print(flush=True, file=out)


# ═══════════════════════════════════════════════════════════════════
#  Progress Bar Prefixes
# ═══════════════════════════════════════════════════════════════════

PFX_READING = "Reading PDFs:    "
PFX_STAMPING = "Stamping PDFs:   "
PFX_COPYING = "Copying PDFs:    "
PFX_MATCHING = "Matching PDFs:   "
PFX_ARCHIVING = "Archiving PDFs:  "


# ═══════════════════════════════════════════════════════════════════
#  String Utilities
# ═══════════════════════════════════════════════════════════════════


def _sanitise(text: str) -> str:
    return _RE_SPACES.sub(" ", _RE_INVALID.sub("", text)).strip()


def safe_filename(name: str) -> str:
    return _sanitise(name)


def _is_company_name(name: str) -> bool:
    parts = name.split()
    return (len(name) == 27 and len(parts) >= 4) or bool(_RE_COMPANY.search(name))


def _format_insured_name(
    name: str,
    *,
    lessor: bool = False,
    has_bcdl_string: bool = False,
    has_bcdl_number: bool = False,
) -> str:
    name = _sanitise(name).title()
    parts = name.split()

    if has_bcdl_string and not has_bcdl_number:
        return name

    if (lessor or not has_bcdl_string) and _is_company_name(name):
        return name

    if len(parts) == 1:
        return name

    return " ".join(parts[1:] + [parts[0]])


def extract_insured_name(
    page_text: str,
    *,
    has_bcdl_string: bool = False,
    has_bcdl_number: bool = False,
) -> str | None:
    lessor = re.search(r"\((?:LESSOR|LSR)\)\s*([^\n]+)", page_text, re.IGNORECASE)
    if lessor:
        return _format_insured_name(
            lessor.group(1).strip(),
            lessor=True,
            has_bcdl_string=has_bcdl_string,
            has_bcdl_number=has_bcdl_number,
        )

    owner = re.search(
        r"(?:Owner\s|Applicant|Name of Insured \(surname followed by given name\(s\)\))\s*\n([^\n]+)",
        page_text,
        re.IGNORECASE,
    )
    if owner:
        return _format_insured_name(
            owner.group(1).strip(),
            has_bcdl_string=has_bcdl_string,
            has_bcdl_number=has_bcdl_number,
        )

    return None


# ═══════════════════════════════════════════════════════════════════
#  Path Utilities
# ═══════════════════════════════════════════════════════════════════


def unique_file_path(path: Path) -> Path:
    base = _RE_COUNTER.sub("", safe_filename(path.stem))
    candidate = path.with_name(f"{base}{path.suffix}")
    counter = 1
    while candidate.exists():
        candidate = path.with_name(f"{base} ({counter}){path.suffix}")
        counter += 1
    return candidate


def _file_key(stem: str) -> str:
    stem = stem.split(" - ", 1)[0] if " - " in stem else stem.split(" ", 1)[0]
    return _RE_INVALID.sub("", stem).upper().strip()


def _extract_filename_timestamp(path: Path) -> str | None:
    m = _RE_FILENAME_TS.search(path.stem)
    return m.group(1) if m else None


# ═══════════════════════════════════════════════════════════════════
#  Excel Mapping
# ═══════════════════════════════════════════════════════════════════


def load_excel_mapping(
    mapping_path: Path | str = Path.cwd() / "config.xlsx",
    sheet_name: str = "config",
) -> FolderMapping:
    mapping_path = Path(mapping_path)
    if not mapping_path.exists():
        return FolderMapping(
            tool_event="ICBC E-Stamp and Copy Tool",
            copy_input_folder=None,
            create_folder_tool_output_folder=None,
            e_stamp_output_folder=None,
        )

    wb = openpyxl.load_workbook(mapping_path)

    sheet_name_resolved = next(
        (s for s in wb.sheetnames if s.casefold() == sheet_name.casefold()),
        None,
    )
    if sheet_name_resolved is None:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )

    ws = wb[sheet_name_resolved]

    def _read_path(row: int) -> Path | None:
        val = ws.cell(row=row, column=2).value
        return Path(val).expanduser() if val else None

    def _read_str(row: int) -> str | None:
        val = ws.cell(row=row, column=2).value
        return str(val).strip() if val else None

    producer_mapping = {
        str(row[0]).upper(): str(row[1])
        for row in ws.iter_rows(min_row=18, values_only=True)
        if row[0] and row[1]
    }

    return FolderMapping(
        tool_event=_read_str(3),
        copy_input_folder=_read_path(7),
        create_folder_tool_output_folder=_read_path(9),
        e_stamp_output_folder=_read_path(13),
        agency_number=_read_str(15),
        producer_mapping=producer_mapping,
    )


# ═══════════════════════════════════════════════════════════════════
#  PDF Scanning — private helpers
# ═══════════════════════════════════════════════════════════════════


def _search(patterns: RegexPatterns, key: str, text: str) -> re.Match[str] | None:
    pat = patterns.get(key)
    return pat.search(text) if pat else None


def _extract_base_fields(
    text: str,
    patterns: RegexPatterns,
) -> tuple[str, str | None, str | None, str | None, str | None, bool]:
    ts = _search(patterns, "timestamp", text)
    cert_rep = _search(patterns, "certificate_replacement", text)
    same_day = _search(patterns, "same_day_re-print", text)

    if not ts and not cert_rep and not same_day:
        raise ValueError("No timestamp found — not an ICBC document")

    raw_timestamp = (
        ts.group(1) if ts else cert_rep.group(1) if cert_rep else same_day.group(1)
    )

    lp = _search(patterns, "license_plate", text)
    bcdl = _search(patterns, "has_bcdl", text)

    return (
        raw_timestamp,
        cert_rep.group(1) if cert_rep else None,
        same_day.group(1) if same_day else None,
        lp.group(1).strip().upper() if lp else None,
        extract_insured_name(
            text,
            has_bcdl_string=bool(bcdl),
            has_bcdl_number=bool(bcdl and bcdl.group(1)),
        ),
        bool(_search(patterns, "temporary_operation_permit", text)),
    )


def _extract_stamping_fields(
    doc: fitz.Document,
    text: str,
    patterns: RegexPatterns,
) -> dict:
    agency = _search(patterns, "agency_number", text)

    customer_copy_pages: list[int] = []
    validation_stamp_coords: list[tuple] = []
    time_of_validation_coords: list[tuple] = []

    for page_num, page in enumerate(doc):
        page_has_customer_copy = False
        for block in page.get_text("blocks"):
            x0, y0, x1, y1, block_text = *block[:4], block[4]
            coords = (page_num, (x0, y0, x1, y1))
            if not page_has_customer_copy and _search(
                patterns, "customer_copy", block_text
            ):
                customer_copy_pages.append(page_num)
                page_has_customer_copy = True
            if _search(patterns, "validation_stamp", block_text):
                validation_stamp_coords.append(coords)
            if _search(patterns, "time_of_validation", block_text):
                time_of_validation_coords.append(coords)

    return {
        "agency_number": agency.group(1).strip() if agency else "UNKNOWN",
        "customer_copy_pages": customer_copy_pages,
        "validation_stamp_coords": validation_stamp_coords,
        "time_of_validation_coords": time_of_validation_coords,
    }


def _extract_copy_fields(
    text: str,
    producer_text: str,
    patterns: RegexPatterns,
) -> dict:
    def flag(key: str) -> bool:
        return bool(_search(patterns, key, text))

    producer = _search(patterns, "producer", producer_text)
    trans = _search(patterns, "transaction_type", text)

    return {
        "producer_name": producer.group(1).upper() if producer else None,
        "transaction_type": trans.group(1).strip().title() if trans else None,
        "storage": flag("storage_policy"),
        "cancellation": flag("cancellation"),
        "special_risk": flag("special_risk_own_damage_policy"),
        "rental": flag("rental_vehicle_policy"),
        "garage": flag("garage_vehicle_certificate"),
        "manuscript": flag("manuscript"),
        "binder": flag("binder"),
    }


# ═══════════════════════════════════════════════════════════════════
#  PDF Scanning — per-file worker (runs inside thread pool)
# ═══════════════════════════════════════════════════════════════════


def _process_one_pdf(
    pdf_path: Path,
    regex_patterns: RegexPatterns,
    page_rects: PageRects,
    stamping_mode: bool,
    copy_mode: bool,
    config_agency_number: str | None,
) -> tuple[Path, str, ICBCDocument | None, str | None]:
    try:
        with fitz.open(pdf_path) as doc:
            if doc.page_count == 0:
                return pdf_path, "non_icbc", None, None

            _text_cache: dict[str | None, str] = {}

            def _page_text(clip_name: str | None = None) -> str:
                if clip_name in _text_cache:
                    return _text_cache[clip_name]
                rect = page_rects.get(clip_name) if clip_name else None
                raw = doc[0].get_text(clip=rect) if rect else doc[0].get_text()
                result = (raw or "").strip()
                _text_cache[clip_name] = result
                return result

            full_text = _page_text()

            if _search(regex_patterns, "payment_plan", full_text) or _search(
                regex_patterns, "payment_plan_receipt", full_text
            ):
                return pdf_path, "payment_plan", None, None

            try:
                (
                    raw_timestamp,
                    certificate_replacement,
                    same_day_reprint,
                    license_plate,
                    insured_name,
                    top,
                ) = _extract_base_fields(full_text, regex_patterns)
            except ValueError:
                return pdf_path, "non_icbc", None, None

            is_replacement = (
                certificate_replacement is not None or same_day_reprint is not None
            )
            effective_timestamp = (
                certificate_replacement or same_day_reprint or raw_timestamp
            )

            document = ICBCDocument(
                path=pdf_path,
                transaction_timestamp=effective_timestamp,
                certificate_replacement=certificate_replacement,
                same_day_reprint=same_day_reprint,
                license_plate=license_plate,
                insured_name=insured_name,
                top=top,
            )

            if stamping_mode:
                for k, v in _extract_stamping_fields(
                    doc, full_text, regex_patterns
                ).items():
                    setattr(document, k, v)
                if is_replacement and config_agency_number:
                    document.agency_number = config_agency_number

            if copy_mode:
                for k, v in _extract_copy_fields(
                    full_text, _page_text("producer"), regex_patterns
                ).items():
                    setattr(document, k, v)

            return pdf_path, "ok", document, None

    except Exception as e:
        return pdf_path, "unreadable", None, str(e)


# ═══════════════════════════════════════════════════════════════════
#  PDF Scanning — public
# ═══════════════════════════════════════════════════════════════════


def scan_icbc_pdfs(
    input_dir: Path | str,
    regex_patterns: RegexPatterns,
    page_rects: PageRects | None = None,
    max_docs: int | None = None,
    stamping_mode: bool = False,
    copy_mode: bool = False,
    config_agency_number: str | None = None,
) -> ScanResult:
    input_dir = Path(input_dir)
    page_rects = page_rects or {}

    pdfs_with_mtime = [(f, f.stat().st_mtime) for f in input_dir.rglob("*.pdf")]
    pdfs = [f for f, _ in sorted(pdfs_with_mtime, key=lambda x: x[1], reverse=True)]
    if max_docs:
        pdfs = pdfs[:max_docs]

    total = len(pdfs)
    bar_size = 10
    _counter = 0
    _lock = threading.Lock()
    _start = time.time()

    def _render(n: int) -> None:
        filled = int(bar_size * n / total) if total else bar_size
        elapsed = time.time() - _start
        remaining = (elapsed / n) * (total - n) if n else 0
        mins, secs = divmod(remaining, 60)
        print(
            f"\r{PFX_READING}[{'█' * filled}{'.' * (bar_size - filled)}] "
            f"{n}/{total} Est wait {int(mins):02}:{secs:03.1f}",
            end="",
            flush=True,
        )

    def _tracked(p: Path):
        nonlocal _counter
        result = _process_one_pdf(
            p,
            regex_patterns,
            page_rects,
            stamping_mode,
            copy_mode,
            config_agency_number,
        )
        with _lock:
            _counter += 1
            _render(_counter)
        return result

    documents: dict[Path, ICBCDocument] = {}
    non_icbc: list[Path] = []
    payment_plans: list[Path] = []
    unreadable: list[Path] = []

    workers = min(8, (os.cpu_count() or 1) * 2)

    if total:
        _render(0)

    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {executor.submit(_tracked, p): p for p in pdfs}
        for future in as_completed(future_map):
            path, category, document, error = future.result()
            if category == "ok":
                documents[path] = document
            elif category == "non_icbc":
                non_icbc.append(path)
            elif category == "payment_plan":
                payment_plans.append(path)
            else:
                unreadable.append(path)

    print(flush=True)

    mtime_order = {p: i for i, p in enumerate(pdfs)}
    documents = dict(sorted(documents.items(), key=lambda kv: mtime_order[kv[0]]))

    return ScanResult(documents, non_icbc, payment_plans, unreadable)


# ═══════════════════════════════════════════════════════════════════
#  Copy PDFs
# ═══════════════════════════════════════════════════════════════════


def copy_pdfs(
    documents: dict[Path, ICBCDocument],
    output_root_dir: Path | str,
    producer_mapping: dict[str, str] | None = None,
    ignore_archive: bool = False,
) -> tuple[list[Path], list[Path]]:
    output_root = Path(output_root_dir)
    prod_map = producer_mapping or {}
    archive_folder = output_root / "_Archive"

    existing_index: dict[str, set[str]] = {}
    for existing in output_root.rglob("*.pdf"):
        if ignore_archive and archive_folder in existing.parents:
            continue
        stem_no_ts = _RE_FILENAME_TS.sub("", existing.stem).strip()
        key = stem_no_ts.split(" - ", 1)[0].strip().lower()
        ts = _extract_filename_timestamp(existing)
        if ts:
            existing_index.setdefault(key, set()).add(ts)

    copied: list[Path] = []
    duplicates: list[Path] = []
    seen: set[tuple[str, str]] = set()

    for src, doc in progressbar(
        list(reversed(list(documents.items()))), prefix=PFX_COPYING, size=10
    ):
        dest_folder = output_root
        if doc.producer_name and doc.producer_name in prod_map:
            dest_folder = output_root / safe_filename(prod_map[doc.producer_name])
        dest_folder.mkdir(parents=True, exist_ok=True)

        base_name = safe_filename(doc.base_name())
        prefix_name = doc.name_prefix
        timestamp = doc.transaction_timestamp
        dedup_key = (prefix_name, timestamp)

        if dedup_key in seen:
            duplicates.append(src)
            continue

        prefix_key = prefix_name.lower()
        if timestamp in existing_index.get(prefix_key, set()):
            duplicates.append(src)
            continue

        dest_name = f"{base_name} [{timestamp}]{src.suffix}"
        dest_file = unique_file_path(dest_folder / dest_name)

        try:
            shutil.copy2(src, dest_file)
            copied.append(dest_file)
            seen.add(dedup_key)
            existing_index.setdefault(prefix_key, set()).add(timestamp)
        except Exception as e:
            print(f"Failed to copy '{src.name}': {e}")

    return copied, duplicates


# ═══════════════════════════════════════════════════════════════════
#  Match PDFs
# ═══════════════════════════════════════════════════════════════════


def _build_match_index(
    subfolder_cache: dict[str, list[Path]],
    root: Path,
) -> dict[str, Path | None]:
    index: dict[str, Path | None] = {}
    for subdir_key, contents in subfolder_cache.items():
        top_level = subdir_key.split("/")[-1]
        is_year = bool(_RE_YEAR.match(top_level))
        for candidate in contents:
            if not candidate.is_file():
                continue
            k = candidate.stem.lower().split(" - ", 1)[0].strip()
            if k in index:
                continue
            index[k] = None if is_year else root / top_level
    return index


def _target_subfolder(
    file: Path,
    root: Path,
    match_index: dict[str, Path | None],
) -> Path:
    if file.parent != root:
        return root
    k = file.stem.split(" - ", 1)[0].strip().lower()
    result = match_index.get(k)
    return result if result is not None else root


def match_pdfs(
    files: list[Path],
    copy_with_no_producer_two: bool,
    root_folder: Path | str,
) -> list[Path] | None:
    if not copy_with_no_producer_two:
        return None

    root = Path(root_folder)

    subfolder_cache: dict[str, list[Path]] = {}
    for subdir in root.rglob("*"):
        if subdir.is_dir() and subdir != root:
            try:
                subfolder_cache[subdir.relative_to(root).as_posix()] = list(
                    subdir.iterdir()
                )
            except PermissionError:
                continue

    match_index = _build_match_index(subfolder_cache, root)

    moved: list[Path] = []
    for file in progressbar(files, prefix=PFX_MATCHING, size=10):
        target = _target_subfolder(file, root, match_index)
        if target == file.parent:
            continue
        target.mkdir(parents=True, exist_ok=True)
        dest = unique_file_path(target / file.name)
        shutil.move(str(file), dest)
        moved.append(dest)

    return moved


# ═══════════════════════════════════════════════════════════════════
#  Auto Archive
# ═══════════════════════════════════════════════════════════════════


def auto_archive(
    root_path: Path | str,
    min_age_years: int = 2,
) -> list[Path] | None:
    root = Path(root_path)
    archive = root / "_Archive"
    archive.mkdir(exist_ok=True)

    cutoff = (datetime.now() - timedelta(days=365 * min_age_years)).date()

    all_pdfs = [f for f in root.rglob("*.pdf") if archive not in f.parents]

    stale = [
        p for p in all_pdfs if datetime.fromtimestamp(p.stat().st_mtime).date() < cutoff
    ]
    if not stale:
        return None

    archived: list[Path] = []
    for pdf in progressbar(stale, prefix=PFX_ARCHIVING, size=10):
        year = time.strftime("%Y", time.localtime(pdf.stat().st_mtime))
        target = archive / year / pdf.relative_to(root).parent
        target.mkdir(parents=True, exist_ok=True)
        dest = unique_file_path(target / pdf.name)
        shutil.move(str(pdf), dest)
        archived.append(dest)

    return archived


# ═══════════════════════════════════════════════════════════════════
#  Reincrement PDFs
# ═══════════════════════════════════════════════════════════════════


def reincrement_pdfs(root_dir: Path | str) -> None:
    root = Path(root_dir)
    if not root.is_dir():
        return

    for folder in sorted([root, *root.rglob("*")], key=lambda f: f.parts, reverse=True):
        if not folder.is_dir():
            continue

        groups: defaultdict[str, list[tuple[int, Path]]] = defaultdict(list)
        for pdf in folder.glob("*.pdf"):
            base = _RE_COUNTER.sub("", safe_filename(pdf.stem))
            num_match = re.search(r"\((\d+)\)$", pdf.stem)
            groups[base].append((int(num_match.group(1)) if num_match else 0, pdf))

        for base, entries in groups.items():
            if len(entries) == 1 and entries[0][0] == 0:
                continue
            for i, (_, pdf) in enumerate(sorted(entries)):
                new_name = f"{base}.pdf" if i == 0 else f"{base} ({i}).pdf"
                new_path = pdf.with_name(new_name)
                if new_path != pdf:
                    pdf.rename(unique_file_path(new_path))

        if folder != root and not any(folder.iterdir()):
            folder.rmdir()


# ═══════════════════════════════════════════════════════════════════
#  Stamping Constants
# ═══════════════════════════════════════════════════════════════════

VALIDATION_STAMP_OFFSET = (-4.25, 23.77, 1.58, 58.95)
TIME_OF_VALIDATION_OFFSET = (0.0, 10.35, 0.0, 40.0)
TIME_STAMP_OFFSET = (0.0, 13.0, 0.0, 0.0)
TIME_OF_VALIDATION_AM_OFFSET = (0.0, -0.6, 0.0, 0.0)
TIME_OF_VALIDATION_PM_OFFSET = (0.0, 21.2, 0.0, 0.0)


# ═══════════════════════════════════════════════════════════════════
#  PDF Stamping Functions
# ═══════════════════════════════════════════════════════════════════


def find_existing_timestamps(
    base_name: str,
    folder_dir: Path | str,
) -> set[str]:
    folder = Path(folder_dir)
    key = _file_key(base_name)
    return {
        ts
        for pdf in folder.rglob("*.pdf")
        if _file_key(pdf.stem) == key and (ts := _extract_filename_timestamp(pdf))
    }


def validation_stamp(
    doc: fitz.Document, document: ICBCDocument, ts_dt: datetime
) -> fitz.Document:
    for page_num, (x0, y0, x1, y1) in document.validation_stamp_coords:
        dx0, dy0, dx1, dy1 = VALIDATION_STAMP_OFFSET
        agency_rect = fitz.Rect(x0 + dx0, y0 + dy0, x1 + dx1, y1 + dy1)
        date_rect = fitz.Rect(
            agency_rect.x0 + TIME_STAMP_OFFSET[0],
            agency_rect.y0 + TIME_STAMP_OFFSET[1],
            agency_rect.x1 + TIME_STAMP_OFFSET[2],
            agency_rect.y1 + TIME_STAMP_OFFSET[3],
        )
        page = doc[page_num]
        page.insert_textbox(
            agency_rect,
            document.agency_number,
            fontname="spacembo",
            fontsize=9,
            align=1,
        )
        page.insert_textbox(
            date_rect,
            ts_dt.strftime("%b %d, %Y"),
            fontname="spacemo",
            fontsize=9,
            align=1,
        )
    return doc


def stamp_time_of_validation(
    doc: fitz.Document, document: ICBCDocument, ts_dt: datetime
) -> fitz.Document:
    am_pm_offset = (
        TIME_OF_VALIDATION_AM_OFFSET
        if ts_dt.hour < 12
        else TIME_OF_VALIDATION_PM_OFFSET
    )
    for page_num, (x0, y0, x1, y1) in document.time_of_validation_coords:
        dx0, dy0, dx1, dy1 = TIME_OF_VALIDATION_OFFSET
        dx0 += am_pm_offset[0]
        dy0 += am_pm_offset[1]
        time_rect = fitz.Rect(x0 + dx0, y0 + dy0, x1 + dx1, y1 + dy1)
        doc[page_num].insert_textbox(
            time_rect, ts_dt.strftime("%I:%M"), fontname="helv", fontsize=6, align=2
        )
    return doc


def save_batch_copy(
    doc: fitz.Document, document: ICBCDocument, output_folder: Path
) -> Path:
    batch_dir = output_folder / "ICBC Batch Copies"
    batch_dir.mkdir(parents=True, exist_ok=True)
    dest = unique_file_path(
        batch_dir / f"{document.base_name()} [{document.transaction_timestamp}].pdf"
    )
    doc.save(dest, garbage=4, deflate=True)
    return dest


def save_customer_copy(
    doc: fitz.Document, document: ICBCDocument, output_folder: Path
) -> Path:
    customer_pages = list(document.customer_copy_pages)
    if document.top and (doc.page_count - 1) not in customer_pages:
        customer_pages.append(doc.page_count - 1)
    pages_to_delete = [i for i in range(doc.page_count) if i not in customer_pages]
    for page_num in reversed(pages_to_delete):
        doc.delete_page(page_num)
    dest = unique_file_path(
        output_folder / f"{document.stamp_name()} (Customer Copy).pdf"
    )
    doc.save(dest, garbage=4, deflate=True)
    return dest
