from pathlib import Path
import fitz
import re
import openpyxl
import shutil
import os
import time
import sys


# --- Clip regions & regex patterns ---
timestamp_location = fitz.Rect(409.979, 63.8488, 576.0, 83.7455)
payment_plan_location = fitz.Rect(425.402, 35.9664, 557.916, 48.3001)
payment_plan_receipt_location = fitz.Rect(
    461.0710144042969, 37.42332458496094, 575.9221801757812, 48.42332077026367
)
producer_location = fitz.Rect(198.0, 761.04, 255.011, 769.977)

timestamp_pattern = re.compile(r"Transaction Timestamp\s*(\d+)")
payment_plan_pattern = re.compile(r"Payment Plan Agreement", re.IGNORECASE)
payment_plan_receipt_pattern = re.compile(r"Payment Plan Receipt")
license_plate_pattern = re.compile(
    r"Licence Plate Number\s*([A-Z0-9\- ]+)", re.IGNORECASE
)
producer_pattern = re.compile(r"-\s*([A-Za-z]+)\s*-", re.IGNORECASE)
transaction_type_pattern = re.compile(r"Transaction Type\s+([A-Z]+)", re.IGNORECASE)
cancellation_pattern = re.compile("Application for Cancellation")
storage_policy_pattern = re.compile(r"Storage Policy")
temporary_operation_permit_pattern = re.compile(
    r"Temporary Operation Permit and Owner’s Certificate of Insurance"
)
rental_vehicle_policy_pattern = re.compile(r"Rental Vehicle Policy")
special_risk_own_damage_policy_pattern = re.compile(r"Special Risk Own Damage Policy")
garage_vehicle_certificate_pattern = re.compile(r"Garage Vehicle Certificate")


# --- Helper functions ---
def progressbar(it, prefix="", size=60, out=sys.stdout):
    count = len(it)
    start = time.time()

    def show(j):
        x = int(size * j / count)
        remaining = ((time.time() - start) / j) * (count - j) if j else 0
        mins, sec = divmod(remaining, 60)
        time_str = f"{int(mins):02}:{sec:03.1f}"
        print(
            f"{prefix}[{'█' * x}{'.' * (size - x)}] {j}/{count} Est wait {time_str}",
            end="\r",
            file=out,
            flush=True,
        )

    if len(it) > 0:
        show(0.1)
        for i, item in enumerate(it):
            yield item
            show(i + 1)
        print(flush=True, file=out)


def reverse_insured_name(name):
    if not name:
        return ""
    name = re.sub(r"\s+", " ", name.strip())
    if name.endswith(("Ltd", "Inc")):
        return name
    name = name.replace(",", "")
    parts = name.split(" ")
    if len(parts) == 1:
        return name
    return " ".join(parts[1:] + [parts[0]])


def search_insured_name(full_text_first_page):
    match = re.search(
        r"(?:Owner\s|Applicant|Name of Insured \(surname followed by given name\(s\)\))\s*\n([^\n]+)",
        full_text_first_page,
        re.IGNORECASE,
    )
    if match:
        name = match.group(1)
        name = re.sub(r"[.:/\\*?\"<>|]", "", name)
        name = re.sub(r"\s+", " ", name).strip().title()
        return name
    return None


def scan_icbc_pdfs(input_dir, max_docs=None):
    """Scan PDFs in a folder and extract metadata."""
    input_dir = Path(input_dir)
    icbc_data = {}
    docs = sorted(
        input_dir.glob("*.pdf"), key=lambda f: f.stat().st_mtime, reverse=True
    )
    if max_docs:
        docs = docs[:max_docs]

    for doc_path in docs:
        try:
            with fitz.open(doc_path) as doc:
                page = doc[0]

                def text(clip=None):
                    return (page.get_text(clip=clip) or "").strip()

                ts_match = timestamp_pattern.search(text(timestamp_location))
                if not ts_match:
                    continue
                timestamp = ts_match.group(1)

                if payment_plan_pattern.search(text(payment_plan_location)):
                    continue
                if payment_plan_receipt_pattern.search(
                    text(payment_plan_receipt_location)
                ):
                    continue

                producer_match = producer_pattern.search(text(producer_location))
                producer_name = (
                    producer_match.group(1).upper() if producer_match else None
                )

                full_text = text()
                license_match = license_plate_pattern.search(full_text)
                license_plate = (
                    license_match.group(1).strip().upper() if license_match else None
                )

                trans_match = transaction_type_pattern.search(full_text)
                transaction_type = (
                    trans_match.group(1).strip().title() if trans_match else None
                )

                insured_name = reverse_insured_name(search_insured_name(full_text))

                top = bool(temporary_operation_permit_pattern.search(full_text))
                storage = bool(storage_policy_pattern.search(full_text))
                cancellation = bool(cancellation_pattern.search(full_text))
                special_risk = bool(
                    special_risk_own_damage_policy_pattern.search(full_text)
                )
                rental = bool(rental_vehicle_policy_pattern.search(full_text))
                garage = bool(garage_vehicle_certificate_pattern.search(full_text))

                icbc_data[doc_path] = {
                    "transaction_timestamp": timestamp,
                    "license_plate": license_plate,
                    "insured_name": insured_name,
                    "producer_name": producer_name,
                    "transaction_type": transaction_type,
                    "top": top,
                    "storage": storage,
                    "cancellation": cancellation,
                    "special_risk": special_risk,
                    "rental": rental,
                    "garage": garage,
                }
        except Exception as e:
            print(f"⚠️  Error processing {doc_path.name}: {e}")
    return icbc_data


# --- Excel loader function ---
def load_producer_mapping(mapping_path):
    """
    Load root folder path and producer-to-folder mapping from Excel.
    Returns (root_folder: Path or None, producer_mapping: dict).
    """
    mapping_path = Path(mapping_path)
    producer_mapping = {}
    root_folder = None

    if mapping_path.exists():
        wb = openpyxl.load_workbook(mapping_path)
        ws = wb.active
        root_cell = ws.cell(row=1, column=2).value
        if root_cell:
            root_folder = Path(root_cell).expanduser()

        for row in ws.iter_rows(min_row=2, values_only=True):
            producer, folder = row
            if producer and folder:
                producer_mapping[str(producer).upper()] = str(folder)

    return root_folder, producer_mapping


# --- Utility functions ---
def safe_filename(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]', "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def unique_file_name(path):
    filename, extension = os.path.splitext(path)
    counter = 1
    while Path(path).is_file():
        path = f"{filename} ({counter}){extension}"
        counter += 1
    return path


def get_base_name(info):
    transaction_timestamp = info.get("transaction_timestamp") or ""
    license_plate = (info.get("license_plate") or "").strip().upper()
    insured_name = (info.get("insured_name") or "").strip()
    insured_name = re.sub(r"[.:/\\*?\"<>|]", "", insured_name)
    insured_name = re.sub(r"\s+", " ", insured_name).strip()
    insured_name = insured_name.title() if insured_name else ""
    transaction_type = (info.get("transaction_type") or "").strip().title()
    top = info.get("top", False)
    storage = info.get("storage", False)
    cancellation = info.get("cancellation", False)
    rental = info.get("rental", False)
    special_risk = info.get("special_risk", False)
    garage = info.get("garage", False)

    if license_plate and license_plate != "NONLIC":
        base_name = license_plate
    elif insured_name:
        base_name = insured_name
    elif transaction_timestamp:
        base_name = transaction_timestamp
    else:
        base_name = "UNKNOWN"

    if top:
        base_name = f"{base_name} TOP"
    elif storage:
        base_name = f"{base_name} Storage Policy"
    elif cancellation:
        base_name = f"{base_name} Cancel"
    elif rental:
        base_name = f"{base_name} Rental Policy"
    elif special_risk:
        base_name = f"{base_name} Special Own Risk Damage"
    elif garage:
        base_name = f"{base_name} Garage Policy"
    elif transaction_type == "Change":
        base_name = f"{base_name} Change"
    elif license_plate == "NONLIC":
        base_name = f"{base_name} Registration"
    return base_name


def pdf_transaction_timestamp(pdf_path):
    try:
        with fitz.open(pdf_path) as doc:
            page = doc[0]
            ts_text = page.get_text(clip=timestamp_location) or ""
            match = timestamp_pattern.search(ts_text)
            if match:
                return match.group(1)
    except Exception as e:
        print(f"⚠️ Error reading {pdf_path.name}: {e}")
    return None


# --- PDF copy function ---
def copy_pdfs(icbc_data, root_folder, producer_mapping=None):
    """Copy scanned PDFs to organized folders."""
    root_folder = Path(root_folder)
    if not root_folder.exists():
        print(f"⚠️ Root folder '{root_folder}' does not exist. Aborting copy.")
        return 0
    producer_mapping = producer_mapping or {}
    copied_count = 0

    # Create a reversed list of items for iteration
    items_to_process = list(reversed(list(icbc_data.items())))

    for path, info in progressbar(items_to_process, prefix="Copying PDFs: ", size=10):
        producer_name = info.get("producer_name")
        if producer_name and producer_name in producer_mapping:
            producer_folder_name = safe_filename(producer_mapping[producer_name])
            subfolder_path = root_folder / producer_folder_name
        else:
            subfolder_path = root_folder

        base_name = get_base_name(info)
        base_name = safe_filename(base_name)
        prefix_name = base_name.split(" ")[0]
        dest_file = subfolder_path / f"{base_name}{path.suffix}"

        # Duplicate check
        duplicate_found = False
        for existing_file in subfolder_path.rglob(f"{prefix_name}*.pdf"):
            existing_ts = pdf_transaction_timestamp(existing_file)
            if existing_ts == info.get("transaction_timestamp"):
                duplicate_found = True
                break

        if not duplicate_found:
            dest_file = Path(unique_file_name(str(dest_file)))
            try:
                shutil.copy2(path, dest_file)
                copied_count += 1
            except Exception as e:
                print(f"⚠️ Failed to copy '{path.name}': {e}")

    return copied_count


# --- Main ---
if __name__ == "__main__":
    input_folder = Path.home() / "Downloads"
    mapping_path = Path.cwd() / "config.xlsx"

    root_folder, producer_mapping = load_producer_mapping(mapping_path)
    scanned_data = scan_icbc_pdfs(input_folder)

    if root_folder:
        copy_pdfs(
            scanned_data,
            root_folder,
            producer_mapping,
        )
