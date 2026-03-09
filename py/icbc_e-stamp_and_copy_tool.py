import fitz
import timeit
import time
import openpyxl
from pathlib import Path
from datetime import datetime
import sys

from utils import (
    _file_key,
    _extract_filename_timestamp,
    progressbar,
    scan_icbc_pdfs,
    load_excel_mapping,
    copy_pdfs,
    match_pdfs,
    auto_archive,
    reincrement_pdfs,
    PFX_STAMPING,
    validation_stamp,
    stamp_time_of_validation,
    save_batch_copy,
    save_customer_copy,
    ICBC_PATTERNS,
    PAGE_RECTS,
)

# ────────────── Constants ────────────── #
DEFAULTS = {
    "number_of_pdfs": 10,
    "copy_with_no_producer_two": True,
    "min_age_to_archive": 1,
    "ignore_archive": False,
}


# ────────────── Shared Utilities ────────────── #


def _countdown(seconds: int) -> None:
    print("Exiting in ", end="", flush=True)
    for i in range(seconds, 0, -1):
        print(f"{i} ", end="", flush=True)
        time.sleep(1)
    print()


def _require_config() -> None:
    mapping_path = Path.cwd() / "config.xlsx"
    if not mapping_path.exists():
        return

    wb = openpyxl.load_workbook(mapping_path, read_only=True)
    if not any(s.casefold() == "config" for s in wb.sheetnames):
        print(f"Missing 'config' sheet in '{mapping_path}'")
        print(f"Available sheets: {wb.sheetnames}")
        print("Please ensure 'config.xlsx' contains a sheet named 'config'.")
        _countdown(7)
        print("Done.")
        sys.exit(1)


# ────────────── ICBC E-Stamp and Copy Tool ────────────── #


def icbc_e_stamp_tool() -> None:
    print("ICBC E-Stamp and Copy Tool\n")
    _require_config()
    start_total = timeit.default_timer()

    # ── Define Desktop stamping folder
    desktop_path = Path.home() / "Desktop"
    if not desktop_path.exists():
        desktop_path = Path.cwd()
    STAMP_OUTPUT_FOLDER = desktop_path / "ICBC E-Stamp Copies"
    STAMP_OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

    # ── Load Excel mapping
    mapping = load_excel_mapping()
    input_folder: Path = Path.home() / "Downloads"
    COPY_OUTPUT_FOLDER: Path = mapping.output_folder
    producer_mapping = mapping.producer_mapping
    copy_mode = bool(COPY_OUTPUT_FOLDER and COPY_OUTPUT_FOLDER.exists())

    # ── Stage 1: Scan PDFs
    scan = scan_icbc_pdfs(
        input_dir=input_folder,
        regex_patterns=ICBC_PATTERNS,
        page_rects=PAGE_RECTS,
        max_docs=DEFAULTS["number_of_pdfs"],
        stamping_mode=True,
        copy_mode=copy_mode,
        config_agency_number=mapping.agency_number,
    )
    total_scanned = len(scan.documents)

    # ── Stage 2: Stamping → Desktop folder
    existing_cache: dict[str, set[str]] = {}
    batch_dir = STAMP_OUTPUT_FOLDER / "ICBC Batch Copies"

    for search_root in (STAMP_OUTPUT_FOLDER, batch_dir):
        if not search_root.exists():
            continue
        for pdf in search_root.rglob("*.pdf"):
            key = _file_key(pdf.stem)
            ts = _extract_filename_timestamp(pdf)
            if ts:
                existing_cache.setdefault(key, set()).add(ts)

    stamped_counter = 0
    for path, document in progressbar(
        list(reversed(list(scan.documents.items()))), prefix=PFX_STAMPING, size=10
    ):
        if not document.transaction_timestamp or not document.validation_stamp_coords:
            continue

        stamp_key = _file_key(document.stamp_name())
        base_key = _file_key(document.base_name())
        existing = existing_cache.get(stamp_key, set()) | existing_cache.get(
            base_key, set()
        )

        if document.transaction_timestamp in existing:
            continue

        ts_dt = datetime.strptime(document.transaction_timestamp, "%Y%m%d%H%M%S")

        try:
            with fitz.open(path) as doc:
                doc = validation_stamp(doc, document, ts_dt)
                doc = stamp_time_of_validation(doc, document, ts_dt)
                save_batch_copy(doc, document, STAMP_OUTPUT_FOLDER)
                save_customer_copy(doc, document, STAMP_OUTPUT_FOLDER)

            existing_cache.setdefault(stamp_key, set()).add(
                document.transaction_timestamp
            )
            existing_cache.setdefault(base_key, set()).add(
                document.transaction_timestamp
            )
            stamped_counter += 1
        except Exception as e:
            print(f"Error processing {path}: {e}")

    if stamped_counter > 0:
        print(
            "\n\033[1m\033[4mStamping complete! ICBC E-Stamp Copies folder is ready now!\033[0m\n"
        )

    # ── Stage 3: Copy → Excel folder
    copied_files = []
    if copy_mode:
        copied_files = copy_pdfs(
            documents=scan.documents,
            output_root_dir=COPY_OUTPUT_FOLDER,
            producer_mapping=producer_mapping,
            ignore_archive=DEFAULTS["ignore_archive"],
        )

        files_without_producer = [
            f for f in copied_files if f.parent == COPY_OUTPUT_FOLDER
        ]
        if files_without_producer:
            match_pdfs(
                files=files_without_producer,
                copy_with_no_producer_two=DEFAULTS["copy_with_no_producer_two"],
                root_folder=COPY_OUTPUT_FOLDER,
            )

        archived_files = auto_archive(
            root_path=COPY_OUTPUT_FOLDER,
            min_age_years=DEFAULTS["min_age_to_archive"],
        )
        if archived_files:
            reincrement_pdfs(root_dir=COPY_OUTPUT_FOLDER)
    else:
        print(
            f"\nNo ICBC Copies folder found — skipping copy step."
            f"\nTo enable copying, set a valid output folder path in B13 of config.xlsx."
        )

    # ── Summary
    elapsed = timeit.default_timer() - start_total
    print(f"\nTotal PDFs scanned: {total_scanned}")
    print(f"Total PDFs stamped: {stamped_counter}")
    print(f"Total PDFs copied:  {len(copied_files)}")
    print(f"Total execution time: {elapsed:.2f} seconds")

    _countdown(3)


# ────────────── Create ICBC Copies Folder Tool ────────────── #


def create_icbc_folder_tool() -> None:
    print("Create ICBC Copies Folder Tool\n")
    _require_config()

    # ── Load config
    mapping = load_excel_mapping()
    input_folder = mapping.copy_input_folder  # B7
    output_folder = mapping.output_folder  # B9
    producer_mapping = mapping.producer_mapping

    # ── Validate folders
    folders_missing = False

    if input_folder and input_folder.exists():
        print(f"Input folder path: {input_folder}")
    else:
        print(f"Input folder '{input_folder}' does not exist.")
        folders_missing = True

    if not output_folder:
        print("Output folder path not set in config.")
        folders_missing = True
    elif not output_folder.parent.exists():
        print(
            f"Parent path '{output_folder.parent}' does not exist. Cannot create output folder."
        )
        folders_missing = True
    else:
        output_folder.mkdir(exist_ok=True)
        print(f"Output folder path: {output_folder}")

    if folders_missing:
        print("Please correct the folder paths in 'config.xlsx'.")
        _countdown(7)
        print("Done.")
        sys.exit(1)

    # ── Scan
    scan = scan_icbc_pdfs(
        input_folder,
        regex_patterns=ICBC_PATTERNS,
        page_rects=PAGE_RECTS,
        max_docs=None,
        copy_mode=True,
    )

    # ── Copy
    copied_files = copy_pdfs(
        documents=scan.documents,
        output_root_dir=output_folder,
        producer_mapping=producer_mapping,
        ignore_archive=DEFAULTS["ignore_archive"],
    )

    # ── Match to producer subfolders
    files_without_producer = [f for f in copied_files if f.parent == output_folder]
    matched_files = match_pdfs(
        files=files_without_producer,
        copy_with_no_producer_two=DEFAULTS["copy_with_no_producer_two"],
        root_folder=output_folder,
    )

    # ── Archive
    archived_files = auto_archive(
        root_path=output_folder,
        min_age_years=DEFAULTS["min_age_to_archive"],
    )
    if archived_files:
        reincrement_pdfs(root_dir=output_folder)

    # ── Log
    log_path = Path.cwd() / "log.txt"
    with open(log_path, "w", encoding="utf-8") as log:
        log.write("=== Create ICBC Copies Folder Tool Summary ===\n")
        log.write("\n")

        if scan.non_icbc:
            log.write("=== Non ICBC PDFs found ===\n")
            log.writelines(f"{p}\n" for p in scan.non_icbc)
            log.write("\n")

        if scan.payment_plans:
            log.write("=== Payment Plan Agreements and Receipts ===\n")
            log.writelines(f"{p}\n" for p in scan.payment_plans)
            log.write("\n")

        if scan.unreadable:
            log.write("=== PDFs that could NOT be opened ===\n")
            log.writelines(f"{p}\n" for p in scan.unreadable)
            log.write("\n")

        if matched_files:
            log.write("=== ICBC PDFs matched to a producer subfolder ===\n")
            log.writelines(f"{p}\n" for p in matched_files)

    print(f"\nLog saved to: {log_path}")
    _countdown(3)


# ────────────── Dispatcher ────────────── #

if __name__ == "__main__":
    _require_config()
    mapping = load_excel_mapping()
    event = (mapping.tool_event or "").strip()

    if event == "Create ICBC Copies Folder Tool":
        create_icbc_folder_tool()
    else:
        if event and event != "ICBC E-Stamp and Copy Tool":
            print(
                f"Unrecognised tool event in B3: '{event}'\n"
                "Defaulting to ICBC E-Stamp and Copy Tool."
            )
        icbc_e_stamp_tool()
